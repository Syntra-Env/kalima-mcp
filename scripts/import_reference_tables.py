"""Import reference datasets (CSV + XLSX) into kalima database.

Run once:  python -X utf8 scripts/import_reference_tables.py

Creates five tables from the Quranic/ CSVs:
  - ref_lemmas        (lemma, lemma_ar, frequency)
  - ref_roots         (root, root_ar, frequency)
  - ref_pos_tags      (pid, pos, pos_ar, pos_en)
  - ref_constituent_tags (pid, tag, tag_ar, tag_en)
  - ref_dependency_rels  (rel_id, rel_en, rel_ar)

Creates one table from the XLSX morph feature sheets:
  - ref_morph_features (category, tag, description_ar, description_en, extra)
"""

import csv
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "kalima.db"
DATASETS = Path(__file__).resolve().parent.parent / "datasets" / "Quranic"

conn = sqlite3.connect(str(DB_PATH))
conn.execute("PRAGMA foreign_keys = ON")


def read_tsv_utf16(path: Path) -> tuple[list[str], list[list[str]]]:
    with open(path, "r", encoding="utf-16") as f:
        reader = csv.reader(f, delimiter="\t")
        header = [h.strip() for h in next(reader)]
        rows = [[c.strip() for c in r] for r in reader if any(c.strip() for c in r)]
    return header, rows


def import_lemmas():
    _, rows = read_tsv_utf16(DATASETS / "CALemmaLexicon.csv")
    conn.execute("DROP TABLE IF EXISTS ref_lemmas")
    conn.execute("""
        CREATE TABLE ref_lemmas (
            lemma TEXT PRIMARY KEY,
            lemma_ar TEXT NOT NULL,
            frequency INTEGER NOT NULL
        )
    """)
    conn.executemany(
        "INSERT INTO ref_lemmas (lemma, lemma_ar, frequency) VALUES (?, ?, ?)",
        [(r[0], r[1], int(r[2])) for r in rows],
    )
    print(f"  ref_lemmas: {len(rows)} rows")


def import_roots():
    _, rows = read_tsv_utf16(DATASETS / "CARootLexicon.csv")
    conn.execute("DROP TABLE IF EXISTS ref_roots")
    conn.execute("""
        CREATE TABLE ref_roots (
            root TEXT PRIMARY KEY,
            root_ar TEXT NOT NULL,
            frequency INTEGER NOT NULL
        )
    """)
    conn.executemany(
        "INSERT INTO ref_roots (root, root_ar, frequency) VALUES (?, ?, ?)",
        [(r[0], r[1], int(r[2])) for r in rows],
    )
    print(f"  ref_roots: {len(rows)} rows")


def import_pos_tags():
    _, rows = read_tsv_utf16(DATASETS / "CAPoS.csv")
    conn.execute("DROP TABLE IF EXISTS ref_pos_tags")
    conn.execute("""
        CREATE TABLE ref_pos_tags (
            pid INTEGER PRIMARY KEY,
            pos TEXT NOT NULL UNIQUE,
            pos_ar TEXT NOT NULL,
            pos_en TEXT NOT NULL
        )
    """)
    conn.executemany(
        "INSERT INTO ref_pos_tags (pid, pos, pos_ar, pos_en) VALUES (?, ?, ?, ?)",
        [(int(r[0]), r[1], r[2], r[3]) for r in rows],
    )
    print(f"  ref_pos_tags: {len(rows)} rows")


def import_constituent_tags():
    _, rows = read_tsv_utf16(DATASETS / "ConstituentsTags.csv")
    conn.execute("DROP TABLE IF EXISTS ref_constituent_tags")
    conn.execute("""
        CREATE TABLE ref_constituent_tags (
            pid INTEGER PRIMARY KEY,
            tag TEXT NOT NULL UNIQUE,
            tag_ar TEXT NOT NULL,
            tag_en TEXT NOT NULL
        )
    """)
    conn.executemany(
        "INSERT INTO ref_constituent_tags (pid, tag, tag_ar, tag_en) VALUES (?, ?, ?, ?)",
        [(int(r[0]), r[1], r[2], r[3]) for r in rows],
    )
    print(f"  ref_constituent_tags: {len(rows)} rows")


def import_dependency_rels():
    _, rows = read_tsv_utf16(DATASETS / "RelLabels.csv")
    conn.execute("DROP TABLE IF EXISTS ref_dependency_rels")
    conn.execute("""
        CREATE TABLE ref_dependency_rels (
            rel_id INTEGER PRIMARY KEY,
            rel_en TEXT NOT NULL,
            rel_ar TEXT NOT NULL
        )
    """)
    conn.executemany(
        "INSERT INTO ref_dependency_rels (rel_id, rel_en, rel_ar) VALUES (?, ?, ?)",
        [(int(r[0]), r[1], r[2]) for r in rows],
    )
    print(f"  ref_dependency_rels: {len(rows)} rows")


def import_morph_features():
    if openpyxl is None:
        print("  ref_morph_features: SKIPPED (openpyxl not installed)")
        return

    wb = openpyxl.load_workbook(str(DATASETS / "CAMorphFeatures.xlsx"))
    conn.execute("DROP TABLE IF EXISTS ref_morph_features")
    conn.execute("""
        CREATE TABLE ref_morph_features (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            tag TEXT NOT NULL,
            description_ar TEXT,
            description_en TEXT,
            extra TEXT
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_ref_morph_category
        ON ref_morph_features(category)
    """)

    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ncols = ws.max_column
        for r in range(2, ws.max_row + 1):
            cells = [ws.cell(r, c).value for c in range(1, ncols + 1)]
            # Skip empty rows
            if not any(cells):
                continue
            # Columns vary per sheet but generally: [id/tag], tag, ar_desc, en_desc, [extra]
            # Normalize: find the tag (non-numeric short string) and descriptions
            str_cells = [str(c).strip() if c is not None else "" for c in cells]

            # Person sheet has no ID column — just tag, ar, en
            if sheet_name == "Person":
                tag = str_cells[0]
                desc_ar = str_cells[1] if len(str_cells) > 1 else ""
                desc_en = str_cells[2] if len(str_cells) > 2 else ""
                extra = ""
            else:
                # Standard: id, tag, ar, en, [extra]
                tag = str_cells[1] if len(str_cells) > 1 else ""
                desc_ar = str_cells[2] if len(str_cells) > 2 else ""
                desc_en = str_cells[3] if len(str_cells) > 3 else ""
                extra = str_cells[4] if len(str_cells) > 4 else ""
                # Clean "None" from extra
                if extra == "None":
                    extra = ""

            conn.execute(
                "INSERT INTO ref_morph_features (category, tag, description_ar, description_en, extra) VALUES (?, ?, ?, ?, ?)",
                (sheet_name, tag, desc_ar, desc_en, extra),
            )
            total += 1

    wb.close()
    print(f"  ref_morph_features: {total} rows across {len(wb.sheetnames)} categories")


def main():
    print("Importing reference tables into kalima database...")
    import_lemmas()
    import_roots()
    import_pos_tags()
    import_constituent_tags()
    import_dependency_rels()
    import_morph_features()
    conn.commit()
    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
